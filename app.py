import pandas as pd
import numpy as np
import folium
import osmnx as ox
import networkx as nx
from geopy import distance
import geopandas as gpd
from geopy.distance import geodesic
from shapely.geometry import Polygon
from shapely.geometry import LineString, Point
from shapely.geometry import box
import random
import math
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')
import streamlit as st

def process_uploaded_files(uploaded_files):
    dataframes = {}
    data = None

    for idx, file in enumerate(uploaded_files):
        df = pd.read_excel(file)

        # Get the filename without extension
        filename_without_extension = file.name.split('.')[0]

        # Assign dataframe to dictionary using filename as key
        dataframes[filename_without_extension] = df

        # Assign specific dataframes
        if idx == 0:
            data = df.copy()

    return dataframes, data


def remove_outliers_district_province(data):
    # Kiểm tra số lượng bản ghi cho mỗi tổ hợp (DistrictName, ProvinceName)
    combo_counts = data.groupby(['DistrictName', 'ProvinceName']).size().reset_index(name='counts')

    # Tìm tổ hợp có số lượng lớn nhất cho mỗi ProvinceName
    max_counts = combo_counts.loc[combo_counts.groupby('ProvinceName')['counts'].idxmax()]

    # Lọc dữ liệu chỉ giữ lại các tổ hợp có số lượng lớn nhất
    filter_district_province = data.merge(max_counts[['DistrictName', 'ProvinceName']], on=['DistrictName', 'ProvinceName'])
    return filter_district_province

def xet_latlng(data):
    min_lat, max_lat = 8.18, 23.39
    min_lon, max_lon = 102.14, 109.46

    # Lọc các dòng không nằm trong khu vực của Việt Nam
    filtered_data = data[
        ~((min_lat <= data['Latitude']) & (data['Latitude'] <= max_lat) &
        (min_lon <= data['Longitude']) & (data['Longitude'] <= max_lon))
    ]

    # Lấy ra những data mà OutletID không nằm trong filtered_data
    unfiltered_outlets = data[~data['OutletID'].isin(filtered_data['OutletID'])]

    return unfiltered_outlets

def tim_tong_oulet_theovung(cleaned_data):
    mean_lat_long = cleaned_data.groupby('WardName')[['Latitude', 'Longitude']].mean().reset_index()
    outlets_per_ward = cleaned_data['WardName'].value_counts().reset_index()
    outlets_per_ward.columns = ['WardName', 'OutletCount']
    ward_data = pd.merge(mean_lat_long, outlets_per_ward, on='WardName')
    return ward_data

def tinh_matdogiapvung(df, place):
    # Tạo một bản đồ với tâm tại khu vực của các xã/thị trấn
    # map_center_lat = df['Latitude'].mean()
    # map_center_lon = df['Longitude'].mean()
    # map_all_wards = folium.Map(location=[map_center_lat, map_center_lon], zoom_start=12)

    # DataFrame để lưu kết quả
    result_data = []

    # Duyệt qua từng xã/thị trấn để lấy ranh giới và thêm vào bản đồ
    for index, row in df.iterrows():
        ward_name = row['WardName']
        latitude = row['Latitude']
        longitude = row['Longitude']
        outlet_count = row['OutletCount']
        
        # Tải ranh giới xã/thị trấn
        gdf = ox.geocode_to_gdf(f"{ward_name}, {place}, Hồ Chí Minh, Vietnam")
        
        if gdf.empty:
            print(f"Không tìm thấy dữ liệu ranh giới cho {ward_name}.")
            continue
        
        # Lấy đường biên (boundary) của ranh giới
        boundary = gdf.geometry.iloc[0].boundary
        
        # # Tạo một màu ngẫu nhiên
        # color = "#{:06x}".format(random.randint(0, 0xFFFFFF))
        
        # # Thêm đường biên vào bản đồ và tô màu
        # folium.features.GeoJson(
        #     boundary,
        #     style_function=lambda feature, color=color: {
        #         'fillColor': color,
        #         'color': color,
        #         'weight': 2,
        #         'fillOpacity': 0.5
        #     },
        #     tooltip=f"{ward_name}\nOutlet Count: {outlet_count}"
        # ).add_to(map_all_wards)

        # Lưu thông tin phường vào DataFrame kết quả
        result_data.append({
            'WardName': ward_name,
            'Boundary': boundary,
            'OutletCount': outlet_count
        })

    # Kiểm tra sự giao nhau của các đường biên
    for i, ward1_data in enumerate(result_data):
        ward1_name = ward1_data['WardName']
        ward1_boundary = ward1_data['Boundary']
        intersecting_wards = []
        
        for j, ward2_data in enumerate(result_data):
            if i != j:
                ward2_name = ward2_data['WardName']
                ward2_boundary = ward2_data['Boundary']
                
                if ward1_boundary.intersects(ward2_boundary):
                    intersecting_wards.append(ward2_name)
        
        # Cập nhật danh sách phường giao nhau vào DataFrame kết quả
        result_data[i]['NumIntersections'] = len(intersecting_wards)
        result_data[i]['IntersectingWards'] = ", ".join(intersecting_wards)

    # Tạo DataFrame từ danh sách kết quả
    result_df = pd.DataFrame(result_data)

    # # Để xem bản đồ trong notebook, bạn có thể dùng:
    # map_all_wards
    
    # Hiển thị kết quả
    return result_df

def cal_reqsr_call_gap(phuong_1, so_visit):
    test = pd.DataFrame(index=[0])
    total_oulet = phuong_1['OutletCount'].sum()
    test['Cal SR'] = int(total_oulet) / (so_visit * 6)
    test['Req SR'] = round(test['Cal SR'])
    test['Call/day'] = round(total_oulet / (test['Req SR'] * 6))
    test['GAP'] = ((test['Call/day'] - so_visit)/so_visit) * 100
    return test

def distance(lat1, lon1, lat2, lon2):
    return geodesic((lat1, lon1), (lat2, lon2)).meters

def tim_group1(df, ward_data, intersecting_df, so_visit, place, threshold):
    start_point = intersecting_df.nsmallest(1, 'NumIntersections')
    start_ward = ward_data[ward_data['WardName'].isin(start_point['WardName'])]
    second_ward = ward_data[ward_data['WardName'].isin(start_point['IntersectingWards'])]
    group_1 = pd.concat([start_ward, second_ward])
    test = cal_reqsr_call_gap(group_1, so_visit)

    priority_updated = False

    while True:
        df = df[~df['WardName'].isin(group_1['WardName'])]

        if len(df) < 0:
            temp_result = group_1
            temp_intersecting_df = intersecting_df
            temp_df = df
            return temp_result, temp_intersecting_df, temp_df
        else:
            ward_name = group_1['WardName'].tail(1).iloc[0]
            temp_data = intersecting_df[intersecting_df['IntersectingWards'].str.contains(ward_name)]
            intersecting_df = tinh_matdogiapvung(df, place)
            
            try:
                temp = temp_data.merge(intersecting_df, on='WardName', how='left', suffixes=('_old', '_new'))
            except Exception as e:
                print(f"Error during merge: {e}")
                temp_result = group_1
                temp_intersecting_df = intersecting_df
                temp_df = df
                return temp_result, temp_intersecting_df, temp_df

            min = temp['NumIntersections_new'].min()
            temp_test = temp[temp['NumIntersections_new'] == min]

            if len(temp_test) > 1:
                df_datagoc = ward_data[ward_data['WardName'] == ward_name]
                df_canxet = ward_data[ward_data['WardName'].isin(temp_test['WardName'])]
                lat_goc = df_datagoc['Latitude'].values[0]
                lon_goc = df_datagoc['Longitude'].values[0]

                distances = []
                for index, row in df_canxet.iterrows():
                    lat_canxet = row['Latitude']
                    lon_canxet = row['Longitude']
                    dist = distance(lat_goc, lon_goc, lat_canxet, lon_canxet)
                    distances.append({
                        'WardName': row['WardName'],
                        'Distance': dist
                    })

                df_distances = pd.DataFrame(distances)
                min_distance_row = df_distances.loc[df_distances['Distance'].idxmin()]
                df_nextward = ward_data[ward_data['WardName'] == min_distance_row['WardName']]
            else:
                df_nextward = ward_data[ward_data['WardName'].isin(temp_test['WardName'])]

            group_1 = pd.concat([group_1, df_nextward])
            test = cal_reqsr_call_gap(group_1, so_visit)

            if 2 <= int(test['Req SR']) <= int(threshold) and -8 <= int(round(test['GAP'])) <= 8:
                temp_result = group_1
                temp_intersecting_df = intersecting_df
                temp_df = df
                priority_updated = True
            elif int(test['Req SR']) <= int(threshold):
                if not priority_updated:
                    temp_result = group_1
                    temp_intersecting_df = intersecting_df
                    temp_df = df

            if int(test['Req SR']) > int(threshold):
                break

    return temp_result, temp_intersecting_df, temp_df

def tim_group2(group_1, df, intersecting_df, ward_data, so_visit, place, threshold):
    group_2 = pd.DataFrame()
    ward_name = group_1['WardName'].tail(1).iloc[0]     
    temp_data = intersecting_df[intersecting_df['IntersectingWards'].str.contains(ward_name)]
    if temp_data.empty:
        print("Chạy lại dùng group 1")
        df = df[~df['WardName'].isin(group_1['WardName'])]
        intersecting_df = tinh_matdogiapvung(df, place)
        temp_result, temp_intersecting_df, temp_df = tim_group1(df, ward_data, intersecting_df, so_visit, place, threshold)
        return temp_result, temp_intersecting_df, temp_df        
    else:
        print("xử lý bình thường")
        df = df[~df['WardName'].isin(group_1['WardName'])]

        intersecting_df = tinh_matdogiapvung(df, place)
        temp = temp_data.merge(intersecting_df, on='WardName', how='left', suffixes=('_old', '_new'))

        min = temp['NumIntersections_new'].min()
        temp_test = temp[temp['NumIntersections_new'] == min]

        if len(temp_test) > 1:
            df_datagoc = ward_data[ward_data['WardName'] == ward_name]
            df_canxet = ward_data[ward_data['WardName'].isin(temp_test['WardName'])]
            # Lấy tọa độ của xã gốc
            lat_goc = df_datagoc['Latitude'].values[0]
            lon_goc = df_datagoc['Longitude'].values[0]

            # Tạo danh sách lưu trữ khoảng cách
            distances = []

            # Duyệt qua các xã trong df_canxet và tính khoảng cách
            for index, row in df_canxet.iterrows():
                lat_canxet = row['Latitude']
                lon_canxet = row['Longitude']
                dist = distance(lat_goc, lon_goc, lat_canxet, lon_canxet)
                distances.append({
                    'WardName': row['WardName'],
                    'Distance': dist
                })

            # Tạo DataFrame từ danh sách khoảng cách
            df_distances = pd.DataFrame(distances)
            min_distance_row = df_distances.loc[df_distances['Distance'].idxmin()]
            df_nextward = ward_data[ward_data['WardName'] == min_distance_row['WardName']]
        else:
            df_nextward = ward_data[ward_data['WardName'].isin(temp_test['WardName'])]

        group_2 = pd.concat([group_2, df_nextward])
        #print(group_2)
        
        # Khởi tạo `temp_result` ngay từ đầu với group_2
        temp_result = group_2.copy()
        temp_intersecting_df = intersecting_df.copy()
        temp_df = df.copy()

        test = cal_reqsr_call_gap(group_2, so_visit)
        priority_updated = False

        while True:
            ward_name = group_2['WardName'].tail(1).iloc[0]
            temp_data = intersecting_df[intersecting_df['IntersectingWards'].str.contains(ward_name)]
            df = df[~df['WardName'].isin(group_2['WardName'])]

            check_soluong = df[~df['WardName'].isin(group_2['WardName'])]
            if len(check_soluong) == 0:
                if not priority_updated and len(group_2) > 0:
                    temp_result = group_2.copy()
                    temp_intersecting_df = None
                    temp_df = None
                break

            intersecting_df = tinh_matdogiapvung(df, place)
            temp = temp_data.merge(intersecting_df, on='WardName', how='left', suffixes=('_old', '_new'))
            
            if temp.empty:
                # Gán giá trị hiện tại cho các biến
                temp_result = group_2.copy()
                temp_intersecting_df = intersecting_df.copy()
                temp_df = df.copy()
                return temp_result, temp_intersecting_df, temp_df
                
            min = temp['NumIntersections_new'].min()
            temp_test = temp[temp['NumIntersections_new'] == min]

            if len(temp_test) > 1:
                df_datagoc = ward_data[ward_data['WardName'] == ward_name]
                df_canxet = ward_data[ward_data['WardName'].isin(temp_test['WardName'])]
                # Lấy tọa độ của xã gốc
                lat_goc = df_datagoc['Latitude'].values[0]
                lon_goc = df_datagoc['Longitude'].values[0]

                # Tạo danh sách lưu trữ khoảng cách
                distances = []

                # Duyệt qua các xã trong df_canxet và tính khoảng cách
                for index, row in df_canxet.iterrows():
                    lat_canxet = row['Latitude']
                    lon_canxet = row['Longitude']
                    dist = distance(lat_goc, lon_goc, lat_canxet, lon_canxet)
                    distances.append({
                        'WardName': row['WardName'],
                        'Distance': dist
                    })

                # Tạo DataFrame từ danh sách khoảng cách
                df_distances = pd.DataFrame(distances)
                min_distance_row = df_distances.loc[df_distances['Distance'].idxmin()]
                df_nextward = ward_data[ward_data['WardName'] == min_distance_row['WardName']]
            else:
                df_nextward = ward_data[ward_data['WardName'].isin(temp_test['WardName'])]
                
            new_group_2 = pd.concat([group_2, df_nextward])
            test = cal_reqsr_call_gap(new_group_2, so_visit)

            if 2 <= int(test['Req SR']) <= int(threshold) and -8 <= int(round(test['GAP'])) <= 8:
                temp_result = new_group_2.copy()
                temp_intersecting_df = intersecting_df.copy()
                temp_df = df.copy()
                priority_updated = True

            elif int(test['Req SR']) <= int(threshold):
                if not priority_updated:
                    temp_result = new_group_2.copy()
                    temp_intersecting_df = intersecting_df.copy()
                    temp_df = df.copy()

            if int(test['Req SR']) > int(threshold):
                break

            group_2 = new_group_2
            #print(group_2)
            #print(test)

        return temp_result, temp_intersecting_df, temp_df

def tao_vong_lap(df, ward_data, intersecting_df, so_visit, place):
    num1 = random.randint(3, 4)
    num2 = random.randint(3, 4)
    temp_result1, temp_intersecting_df1, temp_df1 = tim_group1(df, ward_data, intersecting_df, so_visit, place, num1)
    temp_result2, temp_intersecting_df2, temp_df2 = tim_group2(temp_result1, temp_df1, temp_intersecting_df1, ward_data, so_visit, place, num2)

    test1 = cal_reqsr_call_gap(temp_result1, so_visit)
    i = 1
    temp_result1['ZID'] = i
    test1['ZID'] = i
    test2 = cal_reqsr_call_gap(temp_result2, so_visit)
    i += 1
    temp_result2['ZID'] = i
    test2['ZID'] = i
    tong_gap = pd.concat([test1, test2])
    print("tong_gap" , tong_gap['GAP'].sum())

    tong_df = pd.concat([temp_result1, temp_result2])

    conlai = len(ward_data) - len(tong_df)

    if conlai == 0:
        print("Chỉ có 2 groups")
        return tong_df, tong_gap, conlai

    else:
        print("Nhiều hơn 2 groups")
        while conlai > 0:
            num = random.randint(3, 4)
            # print(num)
            temp_result3, temp_intersecting_df3, temp_df3 = tim_group2(temp_result2, temp_df2, temp_intersecting_df2, ward_data, so_visit, place, num)

            temp_result3['ZID'] = i + 1
            test3 = cal_reqsr_call_gap(temp_result3, so_visit)
            test3['ZID'] = i + 1

            tong_df = pd.concat([tong_df, temp_result3])
            tong_gap = pd.concat([tong_gap, test3])
            conlai = len(ward_data) - len(tong_df)

            if conlai <= 0:
                print(f"Hoàn thành sau lần lặp thứ {i+1}.")
                break

            print(f"Tổng GAP sau lần lặp thứ {i+1}:", tong_gap['GAP'].sum())

            temp_result2 = temp_result3
            temp_df2 = temp_df3
            temp_intersecting_df2 = temp_intersecting_df3
            i += 1

    return tong_df, tong_gap, conlai

def split_dataframe(df, no_call, no_sale):
    # Tính toán số phần tử tối thiểu mỗi nhóm
    min_per_group = no_call * 6
    total_needed = min_per_group * (no_sale - 1)  # Tính cho các nhóm đầu tiên
    remaining = len(df) - total_needed

    # Nếu không đủ số lượng cửa hàng để chia đều, thông báo cảnh báo
    if remaining < min_per_group:
        print(f"Warning: Không đủ số lượng cửa hàng để chia đều cho các nhóm với số lượng yêu cầu. "
              f"Số lượng cửa hàng còn lại ({remaining}) ít hơn yêu cầu tối thiểu ({min_per_group}).")

    # Sắp xếp DataFrame theo Latitude và Longitude
    df_sorted = df.sort_values(by=['Latitude', 'Longitude']).reset_index(drop=True)

    # Tạo các nhóm
    groups = []
    start = 0

    for i in range(no_sale):
        if i < no_sale - 1:
            # Chia nhóm thông thường
            end = start + min_per_group
            group_df = df_sorted.iloc[start:end].copy()
        else:
            # Nhóm cuối cùng chứa tất cả các phần tử còn lại
            group_df = df_sorted.iloc[start:].copy()

        group_df['group'] = i + 1
        groups.append(group_df)
        start = end if i < no_sale - 1 else len(df_sorted)

    return groups

def create_square_map(group_1_df, min_lat, max_lat, min_lon, max_lon):
    # Tính toán tọa độ của tâm hình vuông
    center_lat = (min_lat + max_lat) / 2
    center_lon = (min_lon + max_lon) / 2

    # Tạo bản đồ mới với tọa độ tâm làm trung tâm và phóng to sao cho hình vuông nằm hoàn toàn trong bản đồ
    baby_map = folium.Map(location=[center_lat, center_lon], zoom_start=12)

    # Thêm marker cho mỗi điểm trong group_1_df
    for index, row in group_1_df.iterrows():
        popup_content = f"Order: {index+1}<br>OutletID: {row['OutletID']}<br>OutletName: {row['OutletName']}<br>Latitude: {row['Latitude']}<br>Longitude: {row['Longitude']}"
        folium.Marker(location=[row['Latitude'], row['Longitude']], popup=folium.Popup(popup_content, max_width=300)).add_to(baby_map)

    # Chọn màu ngẫu nhiên từ danh sách các màu
    colors = ['black', 'beige', 'lightblue', 'gray', 'blue', 'darkred', 'lightgreen', 'purple', 'red', 'green', 'lightred', 'darkblue', 'darkpurple', 'cadetblue', 'orange', 'pink', 'lightgray', 'darkgreen', 'pink', 'yellow', 'purple']

    random_color = random.choice(colors)

    # Tạo hình vuông bao quanh các điểm với màu được chọn ngẫu nhiên
    folium.Rectangle(bounds=[(min_lat, min_lon), (max_lat, max_lon)], color=random_color, fill=True, fill_opacity=0.2).add_to(baby_map)

    return baby_map

def find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, group_1_df):
    corners = [(min_lat, min_lon), (min_lat, max_lon), (max_lat, min_lon), (max_lat, max_lon)]

    nearest_points = {}
    
    for corner in corners:
        min_distance = np.inf
        nearest_point = None

        for index, row in group_1_df.iterrows():
            distance = np.sqrt((corner[0] - row['Latitude'])**2 + (corner[1] - row['Longitude'])**2)
            if distance < min_distance:
                min_distance = distance
                nearest_point = row

        nearest_points[corner] = {'point': nearest_point, 'distance': min_distance}

    min_distance_data = min(nearest_points.items(), key=lambda x: x[1]['distance'])

    return min_distance_data

def create_final_df(filtered_df, min_distance_data, no_outlet):
    # Khởi tạo final_df1 với cột dữ liệu tương tự như filtered_df
    final_df1 = pd.DataFrame(columns=filtered_df.columns)

    # Thêm điểm gần nhất vào final_df1
    final_df1 = pd.concat([final_df1, min_distance_data[1]['point'].to_frame().T], ignore_index=True)

    # Tính khoảng cách giữa các điểm trong filtered_df và điểm gần nhất
    filtered_df['distance_to_min'] = filtered_df.apply(lambda row: np.linalg.norm(np.array((row['Latitude'], row['Longitude'])) - np.array(min_distance_data[0])), axis=1)

    # Sắp xếp filtered_df theo khoảng cách tới điểm gần nhất
    filtered_df_sorted = filtered_df.sort_values(by='distance_to_min')

    # Tìm điểm kế tiếp ngắn nhất từ min_distance_data[1]['point']
    current_point = min_distance_data[1]['point']

    # Lặp để thêm điểm cho đến khi final_df1 có đủ 30 điểm
    while len(final_df1) < no_outlet:
        # Lấy điểm kế tiếp ngắn nhất
        next_point_index = filtered_df_sorted.index[0]
        next_point = filtered_df_sorted.iloc[0]

        # Loại bỏ điểm đã chọn khỏi filtered_df_sorted
        filtered_df_sorted = filtered_df_sorted.drop(next_point_index)

        # Nếu điểm kế tiếp không trùng với điểm hiện tại, thêm vào final_df1
        if next_point_index != current_point.name:
            final_df1 = pd.concat([final_df1, pd.DataFrame([next_point], columns=final_df1.columns)], ignore_index=True)

        # Cập nhật điểm hiện tại là điểm kế tiếp đã chọn
        current_point = next_point

    return final_df1

def draw_small_square(final_df1, map):
    # Tính toán tọa độ của hình vuông bao quanh các điểm trong final_df
    min_lat_final1 = final_df1['Latitude'].min()
    max_lat_final1 = final_df1['Latitude'].max()
    min_lon_final1 = final_df1['Longitude'].min()
    max_lon_final1 = final_df1['Longitude'].max()

    # Chọn màu ngẫu nhiên từ danh sách các màu
    colors = ['black', 'beige', 'lightblue', 'gray', 'blue', 'darkred', 'lightgreen', 'purple', 'red', 'green', 'lightred', 'white', 'darkblue', 'darkpurple', 'cadetblue', 'orange', 'pink', 'lightgray', 'darkgreen', 'pink', 'yellow', 'purple']
    random_color = random.choice(colors)

    # Tạo hình vuông bao quanh các điểm trong final_df với màu được chọn ngẫu nhiên
    folium.Rectangle(bounds=[(min_lat_final1, min_lon_final1), (max_lat_final1, max_lon_final1)], color=random_color, fill=True, fill_opacity=0.2).add_to(map)

    return map

def Create_square(cleaned_data, no_oulet, new_map):
    all_data = pd.DataFrame()
    i = 2
    
    while len(cleaned_data) > 0:
        min_lat = cleaned_data['Latitude'].min()
        max_lat = cleaned_data['Latitude'].max()
        min_lon = cleaned_data['Longitude'].min()
        max_lon = cleaned_data['Longitude'].max()

        min_distance_data = find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, cleaned_data)
        cleaned_data = cleaned_data.drop(min_distance_data[1]['point'].name)
        final_df1 = create_final_df(cleaned_data, min_distance_data, no_oulet)
        new_map = draw_small_square(final_df1, new_map)

        final_df1['SRD'] = i
        all_data = pd.concat([all_data, final_df1], ignore_index=True)
        cleaned_data = cleaned_data[~cleaned_data['OutletID'].isin(all_data['OutletID'])]
        print(cleaned_data.info())
        i += 1
        
        if len(cleaned_data) < no_oulet:
            final_df1 = cleaned_data
            final_df1['SRD'] = i
            all_data = pd.concat([all_data, final_df1], ignore_index=True)
            
            return all_data, new_map
    
    return all_data, new_map

################################################################

def calculate_distance(point_coords, filtered_df):
    # Convert point_coords to a Point object
    point = Point(point_coords[::-1])  # Reverse the order of coordinates

    # Calculate distance using Shapely and store in a new column
    filtered_df['distance_to_point'] = filtered_df.apply(
        lambda row: point.distance(Point(row['Longitude'], row['Latitude'])),
        axis=1
    )

    # Find the closest point
    closest_point = filtered_df.loc[filtered_df['distance_to_point'].idxmin()]

    return closest_point['distance_to_point'], closest_point

def draw_optimal_path(visited_points, new_map, G, random_color, group_feature):
    # Extract the last two points from visited_points
    visited_points_df = pd.DataFrame(visited_points.tail(2))
    last_point = (visited_points_df.iloc[-2]['Latitude'], visited_points_df.iloc[-2]['Longitude'])
    final_point = (visited_points_df.iloc[-1]['Latitude'], visited_points_df.iloc[-1]['Longitude'])

    # Find the optimal path using OSMnx
    start_node = ox.distance.nearest_nodes(G, last_point[1], last_point[0])
    destination_node = ox.distance.nearest_nodes(G, final_point[1], final_point[0])

    optimal_path_nodes = ox.shortest_path(G, start_node, destination_node, weight='length')

    if optimal_path_nodes is not None:
        optimal_path_coordinates = [(G.nodes[node]['y'], G.nodes[node]['x']) for node in optimal_path_nodes]

        # Vẽ đường dẫn tối ưu giữa hai điểm cuối cùng
        poly_line = folium.PolyLine(optimal_path_coordinates, color=random_color, weight=2.5, opacity=1)
        poly_line.add_to(group_feature)

        # # Thêm điểm popup cho điểm thứ hai từ cuối
        # last_point_popup = f"Order: {visited_points_df.index[-2] + 1}<br>OutletID: {visited_points_df['OutletID'].iloc[-2]}<br>OutletName: {visited_points_df['OutletName'].iloc[-2]}<br>Latitude: {last_point[0]}<br>Longitude: {last_point[1]}<br>SRD: {visited_points_df['SRD'].iloc[-2]}"
        # last_marker = folium.Marker(location=[last_point[0], last_point[1]], popup=folium.Popup(last_point_popup, max_width=300))
        # last_marker.add_to(group_feature)

        # # Thêm điểm popup cho điểm cuối cùng
        # final_point_popup = f"Order: {visited_points_df.index[-1] + 1}<br>OutletID: {visited_points_df['OutletID'].iloc[-1]}<br>OutletName: {visited_points_df['OutletName'].iloc[-1]}<br>Latitude: {final_point[0]}<br>Longitude: {final_point[1]}<br>SRD: {visited_points_df['SRD'].iloc[-1]}"
        # final_marker = folium.Marker(location=[final_point[0], final_point[1]], popup=folium.Popup(final_point_popup, max_width=300))
        # final_marker.add_to(group_feature)

    else:
        print("No path found between", start_node, "and", destination_node)

    return new_map

def distance(lat1, lon1, lat2, lon2):
    return geodesic((lat1, lon1), (lat2, lon2)).meters

def calculate_distance_between_two_points(point1_coords, point2_coords, graph, filtered_df):
    # Get coordinates of point 1 and point 2
    point1_coords = (point1_coords['Latitude'], point1_coords['Longitude'])
    point2_coords = (point2_coords['Latitude'], point2_coords['Longitude'])

    # Find the closest nodes to point 1 and point 2
    closest_node1 = ox.distance.nearest_nodes(graph, point1_coords[1], point1_coords[0])
    closest_node2 = ox.distance.nearest_nodes(graph, point2_coords[1], point2_coords[0])

    # Find the nearest nodes for all destinations
    destinations_nodes = {}
    for index, row in filtered_df.iterrows():
        dest_node = ox.distance.nearest_nodes(graph, row['Longitude'], row['Latitude'])
        destinations_nodes[index] = dest_node

    # Calculate shortest paths for all destinations
    shortest_paths = {}
    for index, dest_node in destinations_nodes.items():
        try:
            distance1 = nx.shortest_path_length(graph, closest_node1, dest_node, weight='length')
            distance2 = nx.shortest_path_length(graph, closest_node2, dest_node, weight='length')
            shortest_paths[index] = distance1 + distance2
        except nx.NetworkXNoPath:
            shortest_paths[index] = float('inf')  # Assign a large value for unreachable nodes

    return shortest_paths

def find_nearest_point(visited_points, closest_points):
    # Lấy ra hai dòng cuối cùng từ visited_points
    last_two_rows = visited_points.tail(2)

    # Tạo tam giác từ hai điểm cuối cùng trong last_two_rows
    points = pd.DataFrame({'Latitude': [last_two_rows.iloc[0]['Latitude'], last_two_rows.iloc[1]['Latitude']],
                           'Longitude': [last_two_rows.iloc[0]['Longitude'], last_two_rows.iloc[1]['Longitude']]})

    # Khởi tạo các biến để lưu thông tin của tam giác nhỏ nhất
    min_perimeter = float('inf')
    min_perimeter_outlet_info = None

    # Duyệt qua mỗi điểm trong closest_points
    for index, row in closest_points.iterrows():
        # Thêm điểm thứ ba vào tam giác
        point_df = pd.DataFrame({'Latitude': [row['Latitude']], 'Longitude': [row['Longitude']]})
        points = pd.concat([points, point_df], ignore_index=True)
        triangle = Polygon(points)

        # Tính chu vi của tam giác
        perimeter = triangle.length

        # So sánh với chu vi nhỏ nhất đã tìm thấy
        if perimeter < min_perimeter:
            min_perimeter = perimeter
            min_perimeter_outlet_info = {'OutletID': row['OutletID'], 'OutletName': row['OutletName'],
                                         'CustomerAddress': row['CustomerAddress'], 'WardName': row['WardName'],
                                         'DistrictName': row['DistrictName'], 'ProvinceName': row['ProvinceName'],
                                         'Latitude': row['Latitude'], 'Longitude': row['Longitude'],
                                         'ZID': row['ZID'], 'group': row['group'],
                                         'SRD': row['SRD']}            
        # Xóa điểm thứ ba để chuẩn bị cho lần duyệt tiếp theo
        points = points[:-1]

    # Tạo DataFrame từ min_perimeter_outlet_info
    min_perimeter_outlet_df = pd.concat([pd.DataFrame(min_perimeter_outlet_info, index=[0])])

    return min_perimeter_outlet_df

def create_path_2(group_1_df, G):
    if len(group_1_df) == 1:
        print("Chí có 1 Outlet trong group")
        visited_points = group_1_df
    else:
        print("Có nhiều hơn 1 Outlet trong group")
        min_lat = group_1_df['Latitude'].min()
        max_lat = group_1_df['Latitude'].max()
        min_lon = group_1_df['Longitude'].min()
        max_lon = group_1_df['Longitude'].max()

        # Sử dụng hàm để tạo bản đồ
        baby_map = create_square_map(group_1_df, min_lat, max_lat, min_lon, max_lon)

        # Sử dụng để tìm điểm gần góc
        min_distance_data = find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, group_1_df)

        # Lọc data
        filtered_df = group_1_df.drop(min_distance_data[1]['point'].name)

        # Assuming min_distance_data is defined somewhere in your code
        nearest_point_coords = min_distance_data[0]

        # Call the function to calculate distances and find the closest point
        closest_distance, closest_point = calculate_distance(nearest_point_coords, filtered_df)

        # print("Closest distance:", closest_distance)
        # print("Closest point:", closest_point)   

        start_point = min_distance_data[1]['point']
        # Tạo DataFrame rỗng để lưu các điểm đã thăm
        visited_points = pd.DataFrame(columns=['OutletID', 'OutletName', 'CustomerAddress', 'WardName', 'DistrictName', 'ProvinceName', 'Latitude', 'Longitude', 'ZID', 'group', 'SRD'])

        # Thêm start_point vào DataFrame
        visited_points = pd.concat([visited_points, start_point.to_frame().T], ignore_index=True)
        visited_points = pd.concat([visited_points, closest_point.to_frame().T], ignore_index=True)    
        filtered_sorted_df = group_1_df[~group_1_df['OutletID'].isin(visited_points['OutletID'])]

        while not filtered_sorted_df.empty:
            last_row = visited_points.tail(1).iloc[0]
            # Tâm của hình tròn
            center_lat, center_lon = last_row['Latitude'], last_row['Longitude']

            # Bán kính ban đầu của hình tròn
            radius = 300

            # Lặp cho đến khi tìm được ít nhất một điểm hoặc không thể tăng bán kính nữa
            while True:
                # Lọc dữ liệu
                filtered_data = []
                for index, row in filtered_sorted_df.iterrows():
                    point_lat, point_lon = row['Latitude'], row['Longitude']
                    if distance(center_lat, center_lon, point_lat, point_lon) <= radius:
                        filtered_data.append(row)

                # Tạo DataFrame từ dữ liệu lọc
                filtered_df_within_circle = pd.DataFrame(filtered_data)

                # Kiểm tra nếu có ít nhất một điểm trong hình tròn
                if len(filtered_df_within_circle) > 0:
                    break
                
                # Nếu không có điểm nào và bán kính đã tăng lên, tăng bán kính thêm 100m và tiếp tục lặp
                radius += 300

            # In ra filtered_df_within_circle
            print(filtered_df_within_circle)

            last_two_rows = visited_points.tail(2)
            distances_between_two_points = calculate_distance_between_two_points(last_two_rows.iloc[0], last_two_rows.iloc[1], G, filtered_df_within_circle)

            closest_points_indices = sorted(distances_between_two_points, key=distances_between_two_points.get)[:2]
            closest_points = filtered_df_within_circle.loc[closest_points_indices]
            min_perimeter_outlet_df = find_nearest_point(visited_points, closest_points)

            visited_points = pd.concat([visited_points, min_perimeter_outlet_df], ignore_index=True)
            filtered_sorted_df = group_1_df[~group_1_df['OutletID'].isin(visited_points['OutletID'])]
            filtered_sorted_df.info()
      
    return visited_points

def find_nearest_point_1(visited_points, closest_points):
    # Lấy ra hai dòng cuối cùng từ visited_points
    last_two_rows = visited_points.tail(2)

    # Tạo tam giác từ hai điểm cuối cùng trong last_two_rows
    points = pd.DataFrame({'Latitude': [last_two_rows.iloc[0]['Latitude'], last_two_rows.iloc[1]['Latitude']],
                           'Longitude': [last_two_rows.iloc[0]['Longitude'], last_two_rows.iloc[1]['Longitude']]})

    # Khởi tạo các biến để lưu thông tin của tam giác nhỏ nhất
    min_perimeter = float('inf')
    min_perimeter_outlet_info = None

    # Duyệt qua mỗi điểm trong closest_points
    for index, row in closest_points.iterrows():
        # Thêm điểm thứ ba vào tam giác
        point_df = pd.DataFrame({'Latitude': [row['Latitude']], 'Longitude': [row['Longitude']]})
        points = pd.concat([points, point_df], ignore_index=True)
        triangle = Polygon(points)

        # Tính chu vi của tam giác
        perimeter = triangle.length

        # So sánh với chu vi nhỏ nhất đã tìm thấy
        if perimeter < min_perimeter:
            min_perimeter = perimeter
            min_perimeter_outlet_info = {'OutletID': row['OutletID'], 'OutletName': row['OutletName'],
                                         'CustomerAddress': row['CustomerAddress'], 'WardName': row['WardName'],
                                         'DistrictName': row['DistrictName'], 'ProvinceName': row['ProvinceName'],
                                         'Latitude': row['Latitude'], 'Longitude': row['Longitude'],
                                         'group': row['group'], 'SRD': row['SRD']}         
        # Xóa điểm thứ ba để chuẩn bị cho lần duyệt tiếp theo
        points = points[:-1]

    # Tạo DataFrame từ min_perimeter_outlet_info
    min_perimeter_outlet_df = pd.concat([pd.DataFrame(min_perimeter_outlet_info, index=[0])])

    return min_perimeter_outlet_df

def create_path(group_1_df, G):
    if len(group_1_df) == 1:
        print("Chí có 1 Outlet trong group")
        visited_points = group_1_df
    else:
        print("Có nhiều hơn 1 Outlet trong group")
        min_lat = group_1_df['Latitude'].min()
        max_lat = group_1_df['Latitude'].max()
        min_lon = group_1_df['Longitude'].min()
        max_lon = group_1_df['Longitude'].max()

        # Sử dụng hàm để tạo bản đồ
        baby_map = create_square_map(group_1_df, min_lat, max_lat, min_lon, max_lon)

        # Sử dụng để tìm điểm gần góc
        min_distance_data = find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, group_1_df)

        # Lọc data
        filtered_df = group_1_df.drop(min_distance_data[1]['point'].name)

        # Assuming min_distance_data is defined somewhere in your code
        nearest_point_coords = min_distance_data[0]

        # Call the function to calculate distances and find the closest point
        closest_distance, closest_point = calculate_distance(nearest_point_coords, filtered_df)

        # print("Closest distance:", closest_distance)
        # print("Closest point:", closest_point)   

        start_point = min_distance_data[1]['point']
        # Tạo DataFrame rỗng để lưu các điểm đã thăm
        visited_points = pd.DataFrame(columns=['OutletID', 'OutletName', 'CustomerAddress', 'WardName', 'DistrictName', 'ProvinceName', 'Latitude', 'Longitude', 'group', 'SRD'])

        # Thêm start_point vào DataFrame
        visited_points = pd.concat([visited_points, start_point.to_frame().T], ignore_index=True)
        visited_points = pd.concat([visited_points, closest_point.to_frame().T], ignore_index=True)    
        filtered_sorted_df = group_1_df[~group_1_df['OutletID'].isin(visited_points['OutletID'])]

        while not filtered_sorted_df.empty:
            last_row = visited_points.tail(1).iloc[0]
            # Tâm của hình tròn
            center_lat, center_lon = last_row['Latitude'], last_row['Longitude']

            # Bán kính ban đầu của hình tròn
            radius = 300

            # Lặp cho đến khi tìm được ít nhất một điểm hoặc không thể tăng bán kính nữa
            while True:
                # Lọc dữ liệu
                filtered_data = []
                for index, row in filtered_sorted_df.iterrows():
                    point_lat, point_lon = row['Latitude'], row['Longitude']
                    if distance(center_lat, center_lon, point_lat, point_lon) <= radius:
                        filtered_data.append(row)

                # Tạo DataFrame từ dữ liệu lọc
                filtered_df_within_circle = pd.DataFrame(filtered_data)

                # Kiểm tra nếu có ít nhất một điểm trong hình tròn
                if len(filtered_df_within_circle) > 0:
                    break
                
                # Nếu không có điểm nào và bán kính đã tăng lên, tăng bán kính thêm 100m và tiếp tục lặp
                radius += 300

            # In ra filtered_df_within_circle
            print(filtered_df_within_circle)

            last_two_rows = visited_points.tail(2)
            distances_between_two_points = calculate_distance_between_two_points(last_two_rows.iloc[0], last_two_rows.iloc[1], G, filtered_df_within_circle)

            closest_points_indices = sorted(distances_between_two_points, key=distances_between_two_points.get)[:2]
            closest_points = filtered_df_within_circle.loc[closest_points_indices]
            min_perimeter_outlet_df = find_nearest_point_1(visited_points, closest_points)

            visited_points = pd.concat([visited_points, min_perimeter_outlet_df], ignore_index=True)
            filtered_sorted_df = group_1_df[~group_1_df['OutletID'].isin(visited_points['OutletID'])]
            filtered_sorted_df.info()
      
    return visited_points

import base64
import pandas as pd
from io import BytesIO
from openpyxl.workbook import Workbook

def download_excel(dataframe, filename):
    # Tạo một đối tượng Workbook từ openpyxl
    wb = Workbook()
    ws = wb.active
    
    # Ghi tên các cột vào hàng đầu tiên
    for c_idx, col_name in enumerate(dataframe.columns, 1):
        ws.cell(row=1, column=c_idx, value=col_name)
    
    # Ghi dữ liệu từ DataFrame vào worksheet từ hàng thứ hai trở đi
    for r_idx, row in enumerate(dataframe.iterrows(), 2):  # Bắt đầu từ hàng thứ hai
        for c_idx, value in enumerate(row[1], 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Lưu workbook vào đối tượng BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    
    # Lấy nội dung từ đối tượng BytesIO và mã hóa nó thành base64
    excel_binary = excel_buffer.getvalue()
    b64 = base64.b64encode(excel_binary).decode()
    
    # Tạo liên kết để tải xuống file Excel với UTF-8 encoding
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8;base64,{b64}" download="{filename}.xlsx">Download Excel</a>'
    
    return href

def remove_outliers_iqr(data, factor = 1.5):
    q1 = np.percentile(data, 20)
    q3 = np.percentile(data, 70)
    iqr = q3 - q1
    lower_bound = q1 - factor * iqr
    upper_bound = q3 + factor * iqr
    return data[(data >= lower_bound) & (data <= upper_bound)]

def main():
    st.markdown("<h1 style='text-align: center; font-size: 55px;'>Traveling Salesman Problem</h1>", unsafe_allow_html=True)

    # Upload files
    st.header("1. Upload Excel File")

    # Kiểm tra số lượng file đã tải lên
    uploaded_files = st.file_uploader("Upload Excel file", type=["xlsx"], accept_multiple_files=True)

    dataframes = {}
    data = None
    final_df1 = None

    if uploaded_files:
        # data = pd.read_excel("8.PJP QUAN BINH CHANH team CD 25052024 update 1.xlsx")
        # data = pd.read_excel("6. PJP Quận 12_CD.xlsx")
        # data = pd.read_excel("4. PJP THỦ ĐỨC CD.xlsx")
        dataframes, data = process_uploaded_files(uploaded_files)
        so_sale = st.slider("Select number salesman:", 0, 10, 30, 1)
        st.text(f"Selected number: {so_sale}")
        
        if st.button("Run"):
            st.header("2. Result")
            st.text("Đang xử lý")
            
            data['Longitude'] = data['Longitude'].astype(float)
            data['Latitude'] = data['Latitude'].astype(float)
            cleaned_data_1 = remove_outliers_district_province(data)
            cleaned_data = xet_latlng(cleaned_data_1)

            if 'Quận Gò Vấp' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'

            if 'Quận 8' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 06', 'WardName'] = 'Ward 6'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 09', 'WardName'] = 'Ward 9'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 08', 'WardName'] = 'Ward 8'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 05', 'WardName'] = 'Ward 5'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 04', 'WardName'] = 'Ward 4'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 03', 'WardName'] = 'Ward 3'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 02', 'WardName'] = 'Ward 2'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'

            if 'Quận 10' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'
                
            if 'Quận 11' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 03', 'WardName'] = 'Ward 3'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 06', 'WardName'] = 'Ward 6'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 09', 'WardName'] = 'Ward 9'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 08', 'WardName'] = 'Ward 8'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 02', 'WardName'] = 'Ward 2'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 04', 'WardName'] = 'Ward 4'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 07', 'WardName'] = 'Ward 7'
                
            if 'Quận 3' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'

            if 'Quận 4' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'  
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 02', 'WardName'] = 'Ward 2'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 03', 'WardName'] = 'Ward 3'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 04', 'WardName'] = 'Ward 4'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 06', 'WardName'] = 'Ward 6'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 05', 'WardName'] = 'Ward 5'

            if 'Quận 5' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'  
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 05', 'WardName'] = 'Ward 5'

            if 'Quận 6' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 06', 'WardName'] = 'Ward 6'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 05', 'WardName'] = 'Ward 5'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 04', 'WardName'] = 'Ward 4'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 03', 'WardName'] = 'Ward 3'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 02', 'WardName'] = 'Ward 2'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'

            if 'Quận Bình Thạnh' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 02', 'WardName'] = 'Ward 2'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 03', 'WardName'] = 'Ward 3'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 06', 'WardName'] = 'Ward 6'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 05', 'WardName'] = 'Ward 5' 
                
            if 'Quận Phú Nhuận' in cleaned_data['DistrictName'].values:
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 03', 'WardName'] = 'Ward 3' 
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 08', 'WardName'] = 'Ward 8'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 09', 'WardName'] = 'Ward 9'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 04', 'WardName'] = 'Ward 4'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 05', 'WardName'] = 'Ward 5'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 01', 'WardName'] = 'Ward 1'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 02', 'WardName'] = 'Ward 2'
                cleaned_data.loc[cleaned_data['WardName'] == 'Phường 07', 'WardName'] = 'Ward 7'
        
            st.text("Đọc map")
            location = cleaned_data['ProvinceName'].iloc[0] + ' VietNam'
            nwork_type = 'bike'
            G = ox.graph_from_place(location, network_type=nwork_type)
                            
            st.text("Đã đọc xong map - move next step")
            
            so_outlet = cleaned_data['OutletID'].nunique()
            so_visit = round(so_outlet / so_sale / 6)
            trungbinh_1anh = round(so_outlet / so_sale)
            print("So_visit", so_visit)
    
            if so_sale >= 5:
                print("Có nhiều hơn 5 sales")
                place = cleaned_data['DistrictName'].iloc[0]

                ward_data = tim_tong_oulet_theovung(cleaned_data)

                df = ward_data.copy()
                intersecting_df = tinh_matdogiapvung(df, place)
                # print(ward_data)
                print(intersecting_df)
                
                all_tong_df = pd.DataFrame()
                all_tong_gap = pd.DataFrame()

                n = 19
                
                for i in range(1, n + 1):
                    print(i)
                    # Giả sử hàm tao_vong_lap trả về 3 DataFrame
                    tong_df, tong_gap, conlai = tao_vong_lap(df, ward_data, intersecting_df, so_visit, place)
                    
                    # Thêm cột 'stt' vào các DataFrame
                    tong_df['stt'] = i
                    tong_gap['stt'] = i

                    if int(tong_gap['Req SR'].sum()) == so_sale:
                    # Nối các DataFrame hiện tại vào DataFrame tổng thể
                        all_tong_df = pd.concat([all_tong_df, tong_df], ignore_index=True)
                        all_tong_gap = pd.concat([all_tong_gap, tong_gap], ignore_index=True)

                # all_tong_df.to_excel("C:/Users/PHAMT16/Desktop/all_tong_df.xlsx")
                # all_tong_gap.to_excel("C:/Users/PHAMT16/Desktop/all_tong_gap.xlsx")
                
                grouped_df = all_tong_df.groupby('stt').agg({
                    'WardName': lambda x: ', '.join(x),  # Nối các giá trị 'WardName' thành một chuỗi
                    'ZID': lambda x: ', '.join(map(str, x))  # Nối các giá trị 'ZID' thành một chuỗi
                }).reset_index()

                grouped_df.columns = ['stt', 'list_WardName', 'list_group']
                grouped_df = grouped_df.drop_duplicates(subset=['list_WardName', 'list_group'])
                
                filtered_tong_df = all_tong_df[all_tong_df['stt'].isin(grouped_df['stt'])]
                filtered_all_tong_gap = all_tong_gap[all_tong_gap['stt'].isin(grouped_df['stt'])]

                # filtered_tong_df.to_excel("C:/Users/PHAMT16/Desktop/all_tong_df1.xlsx")
                # filtered_all_tong_gap.to_excel("C:/Users/PHAMT16/Desktop/all_tong_gap1.xlsx")
                
                filtered_all_tong_gap['abs_diff'] = abs(filtered_all_tong_gap['Call/day'] - so_visit)

                grouped_df = filtered_all_tong_gap.groupby('stt')['abs_diff'].sum().reset_index()

                grouped_df = grouped_df.rename(columns={'abs_diff': 'sum_abs_diff'})
                min_sum_abs_diff_row = grouped_df.loc[grouped_df['sum_abs_diff'].idxmin()]
                
                min_stt = min_sum_abs_diff_row['stt']

                min_tong_df = filtered_tong_df[filtered_tong_df['stt'] == min_stt]
                min_tong_gap = filtered_all_tong_gap[filtered_all_tong_gap['stt'] == min_stt]
                
                min_tong_df.to_excel("C:/Users/PHAMT16/Desktop/outpt/all_tong_df.xlsx")
                min_tong_gap.to_excel("C:/Users/PHAMT16/Desktop/outpt/all_tong_gap.xlsx")
                
                data_quyhoach = min_tong_df.copy()
                data_quyhoach_gap = min_tong_gap.copy()
                
                result_dff = pd.DataFrame()
                so_nhom = data_quyhoach['ZID'].max()
                for i in range(1, so_nhom + 1):
                    group_i  = data_quyhoach[data_quyhoach['ZID'] == i]
                    data_i = cleaned_data[cleaned_data['WardName'].isin(group_i['WardName'])]
                    no_call = int(data_quyhoach_gap['Call/day'][data_quyhoach_gap['ZID'] == i])
                    no_sale = int(data_quyhoach_gap['Req SR'][data_quyhoach_gap['ZID'] == i])
                    groups = split_dataframe(data_i, no_call, no_sale)
                    combined_df = pd.concat(groups).reset_index(drop=True)
                    combined_df['ZID'] = i
                #     result_dff = pd.concat([result_dff, combined_df])
                # result_dff
                    sovongcanchinh = combined_df['group'].max()
                    data_no_outlier = pd.DataFrame()
                    for i in range(1, sovongcanchinh+1):
                        group_1 = combined_df[combined_df['group'] == i]
                        group_1['Longitude'] = group_1['Longitude'].astype(float)
                        group_1['Latitude'] = group_1['Latitude'].astype(float)

                        cleaned_latitude = remove_outliers_iqr(group_1['Latitude'])
                        cleaned_longitude = remove_outliers_iqr(group_1['Longitude'])

                        # Làm sạch dữ liệu
                        cleaned_data_group1 = group_1[(group_1['Latitude'].isin(cleaned_latitude)) & (group_1['Longitude'].isin(cleaned_longitude))]
                        data_no_outlier = pd.concat([data_no_outlier, cleaned_data_group1])
                #     result_dff = pd.concat([result_dff, data_no_outlier])   
                # result_dff
                    data_outlier = combined_df[~combined_df['OutletID'].isin(data_no_outlier['OutletID'])]
                    data_outlier['new_group'] = None
                    for index_outlier, row_outlier in data_outlier.iterrows():
                        min_distance = float('inf')
                        nearest_group = None
                        
                        lat1, lon1 = row_outlier['Latitude'], row_outlier['Longitude']
                        
                        for index_no_outlier, row_no_outlier in data_no_outlier.iterrows():
                            lat2, lon2 = row_no_outlier['Latitude'], row_no_outlier['Longitude']
                            dist = distance(lat1, lon1, lat2, lon2)
                            
                            # Kiểm tra nếu đây là khoảng cách nhỏ nhất
                            if dist < min_distance:
                                min_distance = dist
                                nearest_group = row_no_outlier['group']
                        
                        data_outlier.at[index_outlier, 'new_group'] = nearest_group
                #     result_dff = pd.concat([result_dff, data_outlier])   
                # result_dff
                    data_outlier = data_outlier.drop(columns=['group'])
                    data_outlier = data_outlier.rename(columns={'new_group': 'group'})
                    new_data = pd.concat([data_no_outlier, data_outlier])
                    result_dff = pd.concat([result_dff, new_data])  
                
                result_dff['RID'] = result_dff['ZID'].astype(str) + "." + result_dff['group'].astype(str)  
                result_dff.to_excel("C:/Users/PHAMT16/Desktop/gomnhom.xlsx")
                print(len(cleaned_data) - len(result_dff))
                
                so_zone = result_dff['ZID'].max()
                data_zone = pd.DataFrame()
                for i in range(1, so_zone + 1):
                    nhom_1 = result_dff[result_dff['ZID'] == i]
                    group_i  = data_quyhoach[data_quyhoach['ZID'] == i]
                    no_call = int(data_quyhoach_gap['Call/day'][data_quyhoach_gap['ZID'] == i])
                    data_nhom = pd.DataFrame()
                    so_nhom = nhom_1['group'].max()
                    for j in range(1, so_nhom + 1):
                        data_i = nhom_1[nhom_1['group'] == j]
                        min_lat = data_i['Latitude'].min()
                        max_lat = data_i['Latitude'].max()
                        min_lon = data_i['Longitude'].min()
                        max_lon = data_i['Longitude'].max()
                        new_map = create_square_map(data_i, min_lat, max_lat, min_lon, max_lon)
                        all_data, new_map = Create_square(data_i, no_call, new_map)
                        data_nhom = pd.concat([data_nhom, all_data])
                    data_zone = pd.concat([data_zone, data_nhom])

                # data_zone['SRID'] = data_zone['ZID'].astype(str) + "." + data_zone['group'].astype(str) + "." + data_zone['SRD'].astype(str)
                # data_zone.to_excel("C:/Users/PHAMT16/Desktop/outpt/data_zone.xlsx")
                
                print("Xếp tuyến")
                
                so_nhom = data_zone['ZID'].max()
                visited_points_list = []

                for i in range(1, so_nhom + 1):
                    df = data_zone[data_zone['ZID'] == i]
                    so_group = df['group'].max()
                    print("so_group: ", so_group)
                    
                    for j in range(1, so_group + 1):
                        data_df = df[df['group'] == j]
                        so_srd = data_df['SRD'].max()
                        print("so_srd: ", so_srd)
                        
                        for k in range(2, so_srd + 1):
                            print(k)
                            group_data = data_df[data_df['SRD'] == k]
                            print(group_data)
                            visited_points_i = create_path_2(group_data, G)
                            visited_points_list.append(visited_points_i)

                thu_danhsach = pd.DataFrame()

                for k, df in enumerate(visited_points_list):
                    df['seq'] = range(1, len(df) + 1)
                    thu_danhsach = pd.concat([thu_danhsach, df], ignore_index=True)
                    
                # thu_danhsach.to_excel("C:/Users/PHAMT16/Desktop/outpt/thu_danhsach.xlsx")

                columns_to_drop = ['Số điện thoại dummy', 'distance_to_min', 'distance_to_point']
                thu_danhsach = thu_danhsach.drop(columns=columns_to_drop)
                
                thu_danhsach['RID'] = thu_danhsach['ZID'].astype(str) + "." + thu_danhsach['group'].astype(str)
                thu_danhsach['SRID'] = thu_danhsach['ZID'].astype(str) + "." + thu_danhsach['group'].astype(str) + "." + thu_danhsach['SRD'].astype(str)
                thu_danhsach['Sequence'] = thu_danhsach['ZID'].astype(str) + "." + thu_danhsach['group'].astype(str) + "." + thu_danhsach['SRD'].astype(str) + "." + thu_danhsach['seq'].astype(str)
                
                with pd.ExcelWriter('C:/Users/PHAMT16/Desktop/outpt/Output.xlsx', engine='openpyxl') as writer:  
                    thu_danhsach.to_excel(writer, sheet_name='result', index=False)
                    min_tong_df.to_excel(writer, sheet_name='group_list', index=False)
                    min_tong_gap.to_excel(writer, sheet_name='sale_gap', index=False)
                    
                href_csv = download_excel(thu_danhsach, "thu_danhsach")
                st.markdown(href_csv, unsafe_allow_html=True) 
                
                href_csv1 = download_excel(min_tong_df, "min_tong_df")
                st.markdown(href_csv1, unsafe_allow_html=True)    
                
                href_csv2 = download_excel(min_tong_gap, "min_tong_gap")
                st.markdown(href_csv2, unsafe_allow_html=True)          
                    
            else:
                print("Có ít hơn 5 sales")
                groups = split_dataframe(cleaned_data, so_visit, so_sale)
                combined_df = pd.concat(groups).reset_index(drop=True)
                
                # combined_df.to_excel("C:/Users/PHAMT16/Desktop/combined_df.xlsx")
                
                result_dff = pd.DataFrame()
                sovongcanchinh = combined_df['group'].max()
                data_no_outlier = pd.DataFrame()
                for i in range(1, sovongcanchinh+1):
                    group_1 = combined_df[combined_df['group'] == i]
                    group_1['Longitude'] = group_1['Longitude'].astype(float)
                    group_1['Latitude'] = group_1['Latitude'].astype(float)
                    cleaned_latitude = remove_outliers_iqr(group_1['Latitude'])
                    cleaned_longitude = remove_outliers_iqr(group_1['Longitude'])
                    cleaned_data_group1 = group_1[(group_1['Latitude'].isin(cleaned_latitude)) & (group_1['Longitude'].isin(cleaned_longitude))]
                    data_no_outlier = pd.concat([data_no_outlier, cleaned_data_group1])
                    data_outlier = combined_df[~combined_df['OutletID'].isin(data_no_outlier['OutletID'])]
                #     result_dff = pd.concat([result_dff, data_outlier])   
                # result_dff
                    data_outlier['new_group'] = None
                    for index_outlier, row_outlier in data_outlier.iterrows():
                        min_distance = float('inf')
                        nearest_group = None
                        
                        lat1, lon1 = row_outlier['Latitude'], row_outlier['Longitude']
                        
                        for index_no_outlier, row_no_outlier in data_no_outlier.iterrows():
                            lat2, lon2 = row_no_outlier['Latitude'], row_no_outlier['Longitude']
                            dist = distance(lat1, lon1, lat2, lon2)
                            
                            # Kiểm tra nếu đây là khoảng cách nhỏ nhất
                            if dist < min_distance:
                                min_distance = dist
                                nearest_group = row_no_outlier['group']
                        
                        data_outlier.at[index_outlier, 'new_group'] = nearest_group
                #     result_dff = pd.concat([result_dff, data_outlier])   
                # result_dff
                    data_outlier = data_outlier.drop(columns=['group'])
                    data_outlier = data_outlier.rename(columns={'new_group': 'group'})
                    result_dff = pd.concat([data_no_outlier, data_outlier])
                
                # result_dff.to_excel("C:/Users/PHAMT16/Desktop/combined_df.xlsx")
                
                so_nhom = result_dff['group'].max()
                data_zone = pd.DataFrame()
                for i in range(1, so_nhom + 1):
                    data_i = result_dff[result_dff['group'] == i]
                    min_lat = data_i['Latitude'].min()
                    max_lat = data_i['Latitude'].max()
                    min_lon = data_i['Longitude'].min()
                    max_lon = data_i['Longitude'].max()
                    new_map = create_square_map(data_i, min_lat, max_lat, min_lon, max_lon)
                    all_data, new_map = Create_square(data_i, so_visit, new_map)
                    data_zone = pd.concat([data_zone, all_data])
                
                # data_zone.to_excel("C:/Users/PHAMT16/Desktop/combined_df.xlsx")
                
                print("Xếp tuyến")
                
                so_nhom = data_zone['group'].max()
                visited_points_list = []
                for i in range(1, so_nhom + 1): 
                    data_df = data_zone[data_zone['group'] == i]
                    so_srd = data_df['SRD'].max()
                    print("so_srd: ", so_srd)
                    
                    for j in range(2, so_srd + 1):
                        group_data = data_df[data_df['SRD'] == j]
                        visited_points_i = create_path(group_data, G)
                        visited_points_list.append(visited_points_i)
                
                thu_danhsach = pd.DataFrame()
                for k, df in enumerate(visited_points_list):
                    df['seq'] = range(1, len(df) + 1)
                    thu_danhsach = pd.concat([thu_danhsach, df], ignore_index=True)
                
                columns_to_drop = ['Số điện thoại dummy', 'distance_to_min', 'distance_to_point']
                thu_danhsach = thu_danhsach.drop(columns=columns_to_drop)
                
                thu_danhsach = thu_danhsach.rename(columns={'group': 'ZID'})
                thu_danhsach['SRID'] = thu_danhsach['ZID'].astype(str) + "." + thu_danhsach['SRD'].astype(str)
                thu_danhsach['Sequence'] = thu_danhsach['ZID'].astype(str) + "." + thu_danhsach['SRD'].astype(str) + "." + thu_danhsach['seq'].astype(str)
                
                thu_danhsach.to_excel("C:/Users/PHAMT16/Desktop/thu_danhsach.xlsx")
                href_csv = download_excel(thu_danhsach, "thu_danhsach")
                st.markdown(href_csv, unsafe_allow_html=True) 

if __name__ == '__main__':
    main()